import requests, time, unicodedata, json
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APP_KEY    = '1576318757013'
APP_SECRET = '2d96d060ad7abe5dd82c4d214f3a9de8'
OMIE_CLIENTES   = 'https://app.omie.com.br/api/v1/geral/clientes/'
OMIE_CONTAS     = 'https://app.omie.com.br/api/v1/financas/contareceber/'
OMIE_VENDEDORES = 'https://app.omie.com.br/api/v1/geral/vendedores/'
RPP = 100

def norm(s):
    if not s: return ''
    s = unicodedata.normalize('NFD', str(s))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower().strip()

TAGS_REGIAO = {
    'regiao abc', 'regiao alphaville', 'regiao litoral paulista',
    'regiao oeste e sul', 'regiao sorocaba', 'sao paulo',
    'zona leste', 'zona norte', 'guarulhos', 'litoral sul',
    'alagoas', 'belo horizonte', 'brasilia', 'curitiba',
    'rio de janeiro', 'porto alegre', 'fortaleza', 'recife',
    'salvador', 'manaus', 'belem', 'goiania', 'florianopolis',
    'campo grande', 'maceio', 'natal', 'teresina',
    'santa catarina', 'parana', 'noroeste paulista',
    'interior sp', 'grande sp', 'litoral', 'sul', 'minas gerais'
}

TAGS_MODELO = {
    'associados pro', 'associados', 'design',
    'cessao de direitos', 'associado pro', 'associado'
}

def extrair_caracteristicas(caracteristicas):
    dados = {'diretora_regional': '', 'regiao': '', 'status': '', 'modelo': '', 'gestora': ''}
    if not caracteristicas: return dados
    for c in caracteristicas:
        campo = norm(c.get('campo', ''))
        conteudo = str(c.get('conteudo', '')).strip()
        if any(x in campo for x in ['diretor', 'franquia', 'regional']):
            if 'regiao' not in campo:
                dados['diretora_regional'] = conteudo
        if 'gestor' in campo: dados['gestora'] = conteudo
        elif 'regiao' in campo: dados['regiao'] = conteudo
        elif 'status' in campo: dados['status'] = conteudo
        elif 'modelo' in campo: dados['modelo'] = conteudo
    return dados

def invalido(c):
    cnpj = c.get('cnpj_cpf', '') or ''
    nome = c.get('razao_social', '') or ''
    return (not cnpj or cnpj == '000.000.000-00'
            or not nome.strip() or nome == 'Cliente Consumidor')

def omie_post(url, body):
    for i in range(3):
        try:
            r = requests.post(url, json=body, timeout=30)
            return r.json()
        except Exception:
            if i == 2: raise
            time.sleep(2 ** i)

def calcular_dias(data_br):
    if not data_br: return 0
    try:
        d, m, y = data_br.split('/')
        venc = date(int(y), int(m), int(d))
        return max(0, (date.today() - venc).days)
    except:
        return 0

print('=' * 60)
print('CLUB&CASA - Gerando relatorio de inadimplentes')
print('=' * 60)

# ----------------------------------------------------------------
# [1/4] LOJISTAS - removido filtro de tag 'cliente' que excluía
#        a maioria dos lojistas cadastrados
# ----------------------------------------------------------------
print('\n[1/4] Buscando lojistas no Omie...')
p, tp = 1, 1
lojistas = {}
while p <= tp:
    r = omie_post(OMIE_CLIENTES, {
        'call': 'ListarClientes', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
        'param': [{'pagina': p, 'registros_por_pagina': RPP, 'apenas_importado_api': 'N', 'exibir_caracteristicas': 'S'}]
    })
    if not r or not r.get('clientes_cadastro'): break
    tp = r.get('total_de_paginas', 1)

    for c in r['clientes_cadastro']:
        if invalido(c): continue

        # CORREÇÃO 1: removido o filtro 'cliente' not in nd_tags
        # que excluía lojistas sem essa tag específica
        raw_tags = [t.get('tag', '') for t in c.get('tags', [])]
        nd_tags  = [norm(t) for t in raw_tags]

        # Só exclui se explicitamente marcado como fornecedor (sem ser cliente)
        if 'fornecedor' in nd_tags and 'cliente' not in nd_tags:
            continue

        caracs = extrair_caracteristicas(c.get('caracteristicas', []))

        regiao_tag    = next((raw_tags[i] for i, t in enumerate(nd_tags) if t in TAGS_REGIAO), '')
        modelo_tag    = next((raw_tags[i] for i, t in enumerate(nd_tags) if t in TAGS_MODELO), '')
        distrato_tag  = 'distrato' in nd_tags

        regiao_final  = caracs['regiao']  if caracs['regiao']  else regiao_tag
        modelo_final  = caracs['modelo']  if caracs['modelo']  else modelo_tag
        distrato_final = 'Sim' if (norm(caracs['status']) == 'distrato' or distrato_tag) else 'Nao'

        email_raw = c.get('email', '') or ''
        codigo    = c.get('codigo_cliente_omie')

        lojistas[str(codigo)] = {
            'codigo':            codigo,
            'razao_social':      c.get('razao_social', ''),
            'nome_fantasia':     c.get('nome_fantasia', '') or '',
            'cnpj_cpf':          c.get('cnpj_cpf', '') or '',
            'email':             email_raw.split(',')[0].strip(),
            'telefone':          c.get('telefone1_numero', '') or '',
            'cidade':            c.get('cidade', '') or '',
            'estado':            c.get('estado', '') or '',
            'regiao':            regiao_final if regiao_final else 'Sem Regiao',
            'diretora_regional': caracs['diretora_regional'] if caracs['diretora_regional'] else 'Nao Informada',
            'gestora_carac':     caracs['gestora'] if caracs['gestora'] else '',
            'modelo':            modelo_final if modelo_final else 'Sem Modelo',
            'distrato':          distrato_final,
            'todas_tags':        ', '.join(raw_tags),
            'codigo_vendedor':   str((c.get('recomendacoes') or {}).get('codigo_vendedor', '') or ''),
        }

    print(f'  Pag {p:3}/{tp} - {len(lojistas)} lojistas')
    p += 1
    time.sleep(0.3)

# ----------------------------------------------------------------
# [2/4] GESTORAS
# ----------------------------------------------------------------
print('\n[2/4] Buscando gestoras na API Omie...')
vendedores = {}
try:
    p, tp = 1, 1
    while p <= tp:
        r = omie_post(OMIE_VENDEDORES, {
            'call': 'ListarVendedores', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
            'param': [{'pagina': p, 'registros_por_pagina': 50}]
        })
        if not r or not r.get('cadastro'): break
        tp = r.get('total_de_paginas', 1)
        for v in r['cadastro']:
            cod  = str(v.get('codigo', ''))
            nome = v.get('nome', '') or ''
            if cod and nome:
                vendedores[cod] = nome
        p += 1
        time.sleep(0.3)
    print(f'  {len(vendedores)} gestoras encontradas')
except Exception as e:
    print(f'  Erro: {e}')

# ----------------------------------------------------------------
# [3/4] TÍTULOS ATRASADOS
# CORREÇÃO 2: em vez de break no faultstring, apenas loga e avança
#             para não parar na página 1 por erro pontual da API
# ----------------------------------------------------------------
print('\n[3/4] Buscando titulos ATRASADOS...')
p, tp = 1, 1
titulos   = {}
total_tit = 0
erros_consecutivos = 0

while p <= tp:
    r = omie_post(OMIE_CONTAS, {
        'call': 'ListarContasReceber', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
        'param': [{'pagina': p, 'registros_por_pagina': RPP, 'filtrar_por_status': 'ATRASADO'}]
    })

    # Sem resposta alguma: para
    if not r:
        print(f'  Sem resposta na pag {p}, encerrando.')
        break

    # Erro da API: loga, aguarda e tenta a próxima página
    if r.get('faultstring'):
        erros_consecutivos += 1
        print(f'  Aviso pag {p}: {r["faultstring"]}')
        if erros_consecutivos >= 5:
            print('  Muitos erros consecutivos, encerrando.')
            break
        p += 1
        time.sleep(1.5)
        continue

    erros_consecutivos = 0
    tp   = r.get('total_de_paginas', 1)
    regs = r.get('conta_receber_cadastro') or r.get('lista_contareceber') or []

    for t in regs:
        cod = str(t.get('codigo_cliente_fornecedor', ''))
        if not cod: continue
        if cod not in titulos:
            titulos[cod] = []
        titulos[cod].append(t)
        total_tit += 1

    print(f'  Pag {p:3}/{tp} - {total_tit} titulos | {len(titulos)} clientes')
    p += 1
    time.sleep(0.6)

# ----------------------------------------------------------------
# [4/4] CRUZAMENTO
# ----------------------------------------------------------------
print('\n[4/4] Cruzando dados...')
rows = []
for cod, tits in titulos.items():
    cl = lojistas.get(cod)
    if not cl: continue

    val = round(sum(t.get('valor_documento', 0) for t in tits), 2)

    def sort_key(t):
        dv = t.get('dDtVenc') or '99/99/9999'
        try:
            d, m, y = dv.split('/')
            return (int(y), int(m), int(d))
        except:
            return (9999, 99, 99)

    tits_sorted    = sorted(tits, key=sort_key)
    mais_antigo_dv = tits_sorted[0].get('dDtVenc', '')

    gestora_nome = cl['gestora_carac']
    if not gestora_nome and cl['codigo_vendedor']:
        gestora_nome = vendedores.get(cl['codigo_vendedor'], cl['codigo_vendedor'])

    rows.append({
        'codigo_cliente':         cl['codigo'],
        'Razao Social':           cl['razao_social'],
        'Nome Fantasia':          cl['nome_fantasia'],
        'CNPJ/CPF':               cl['cnpj_cpf'],
        'Email':                  cl['email'],
        'Telefone':               cl['telefone'],
        'Cidade':                 cl['cidade'],
        'Estado':                 cl['estado'],
        'Regiao':                 cl['regiao'],
        'Diretora Regional':      cl['diretora_regional'],
        'Modelo de Negocio':      cl['modelo'],
        'Distrato':               cl['distrato'],
        'Gestora':                gestora_nome if gestora_nome else '',
        'Tags':                   cl['todas_tags'],
        'Qtd Titulos':            len(tits),
        'Valor Atrasado (R$)':    val,
        'Vencimento Mais Antigo': mais_antigo_dv,
        'Dias em Atraso':         calcular_dias(mais_antigo_dv),
        'titulos': [{
            'numero_documento': t.get('numero_documento_fiscal') or t.get('numero_documento') or t.get('cNumParcela') or t.get('nIdTitulo') or 'N/I',
            'data_vencimento':  t.get('dDtVenc') or t.get('data_vencimento') or '',
            'valor_documento':  t.get('valor_documento') or t.get('nValParcela') or 0,
            'dias_atraso':      calcular_dias(t.get('dDtVenc') or t.get('data_vencimento') or ''),
        } for t in tits_sorted],
    })

rows.sort(key=lambda x: x['Valor Atrasado (R$)'], reverse=True)

# ----------------------------------------------------------------
# SALVAR JSON
# ----------------------------------------------------------------
print('\nSalvando inadimplentes.json...')
with open('inadimplentes.json', 'w', encoding='utf-8') as f:
    json.dump({
        'gerado_em':      datetime.now().strftime('%d/%m/%Y %H:%M'),
        'total_clientes': len(rows),
        'total_titulos':  total_tit,
        'valor_total':    round(sum(r['Valor Atrasado (R$)'] for r in rows), 2),
        'inadimplentes':  rows
    }, f, ensure_ascii=False, indent=2)

# ----------------------------------------------------------------
# SALVAR XLSX
# ----------------------------------------------------------------
print('Salvando inadimplentes-clubcasa.xlsx...')

wb = Workbook()

roxo       = PatternFill(start_color='5B52E8', end_color='5B52E8', fill_type='solid')
roxo_claro = PatternFill(start_color='F4F6FB', end_color='F4F6FB', fill_type='solid')
branco     = Font(color='FFFFFF', bold=True, size=12)
preto_bold = Font(bold=True, size=11)
borda = Border(
    left=Side(style='thin',   color='E2E6F0'),
    right=Side(style='thin',  color='E2E6F0'),
    top=Side(style='thin',    color='E2E6F0'),
    bottom=Side(style='thin', color='E2E6F0')
)

total_valor    = sum(r['Valor Atrasado (R$)'] for r in rows)
total_distrato = sum(r['Valor Atrasado (R$)'] for r in rows if r['Distrato'] == 'Sim')

ws_resumo = wb.active
ws_resumo.title = 'Resumo'

ws_resumo['A1'] = 'CLUB&CASA DESIGN'
ws_resumo['A1'].font = Font(bold=True, size=16, color='5B52E8')
ws_resumo.merge_cells('A1:B1')

ws_resumo['A2'] = 'Relatorio de Inadimplencia'
ws_resumo['A2'].font = Font(size=12, color='6B7080')
ws_resumo.merge_cells('A2:B2')

ws_resumo['A3'] = 'Gerado em'
ws_resumo['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')

ws_resumo['A5'] = 'Metrica'
ws_resumo['B5'] = 'Valor'
for cell in [ws_resumo['A5'], ws_resumo['B5']]:
    cell.fill = roxo
    cell.font = branco
    cell.alignment = Alignment(horizontal='center')

def fmt_brl(v):
    return f'R$ {v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

dados_resumo = [
    ('Total de Lojistas Inadimplentes', len(rows)),
    ('Total de Titulos em Aberto',      total_tit),
    ('Valor Total Atrasado',            fmt_brl(total_valor)),
    ('Valor em Distrato',               fmt_brl(total_distrato)),
    ('Ticket Medio',                    fmt_brl(total_valor / len(rows)) if rows else 'R$ 0,00'),
]

for i, (k, v) in enumerate(dados_resumo, start=6):
    ws_resumo[f'A{i}'] = k
    ws_resumo[f'B{i}'] = v
    ws_resumo[f'A{i}'].font = preto_bold
    if i % 2 == 0:
        ws_resumo[f'A{i}'].fill = roxo_claro
        ws_resumo[f'B{i}'].fill = roxo_claro

ws_resumo.column_dimensions['A'].width = 35
ws_resumo.column_dimensions['B'].width = 25

ws = wb.create_sheet('Inadimplentes')

headers = [
    'Codigo Omie', 'Razao Social', 'Nome Fantasia', 'CNPJ/CPF', 'Email',
    'Telefone', 'Cidade', 'Estado', 'Regiao', 'Diretora Regional',
    'Modelo de Negocio', 'Distrato', 'Gestora', 'Tags',
    'Qtd Titulos', 'Valor Atrasado (R$)', 'Vencimento Mais Antigo', 'Dias em Atraso'
]

for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = roxo
    cell.font = branco
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = borda

for i, r in enumerate(rows, start=2):
    valores = [
        r['codigo_cliente'], r['Razao Social'], r['Nome Fantasia'], r['CNPJ/CPF'], r['Email'],
        r['Telefone'], r['Cidade'], r['Estado'], r['Regiao'], r['Diretora Regional'],
        r['Modelo de Negocio'], r['Distrato'], r['Gestora'], r['Tags'],
        r['Qtd Titulos'], r['Valor Atrasado (R$)'], r['Vencimento Mais Antigo'], r['Dias em Atraso']
    ]
    for col, v in enumerate(valores, start=1):
        cell = ws.cell(row=i, column=col, value=v)
        cell.border = borda
        if col == 16:
            cell.number_format = 'R$ #,##0.00'
        if i % 2 == 0:
            cell.fill = roxo_claro

larguras = [12, 35, 30, 18, 30, 15, 20, 8, 22, 20, 20, 10, 22, 30, 10, 18, 18, 12]
for i, w in enumerate(larguras, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A2'
wb.save('inadimplentes-clubcasa.xlsx')

print(f'\n{"=" * 60}')
print(f'CONCLUIDO!')
print(f'Total de lojistas: {len(rows)}')
print(f'Total de titulos:  {total_tit}')
print(f'Valor total:       R$ {total_valor:,.2f}')
print(f'Valor em distrato: R$ {total_distrato:,.2f}')