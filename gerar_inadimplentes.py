import requests, time, unicodedata, json, html as html_lib
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APP_KEY    = '1576318757013'
APP_SECRET = '2d96d060ad7abe5dd82c4d214f3a9de8'
OMIE_CLIENTES   = 'https://app.omie.com.br/api/v1/geral/clientes/'
OMIE_CONTAS     = 'https://app.omie.com.br/api/v1/financas/contareceber/'
OMIE_VENDEDORES = 'https://app.omie.com.br/api/v1/geral/vendedores/'
OMIE_CATEGORIAS = 'https://app.omie.com.br/api/v1/geral/categorias/'
OMIE_EXTRATO    = 'https://app.omie.com.br/api/v1/financas/extrato/'
RPP = 100

CONTAS_RECEBIMENTO = [
    (8453130413, 'CEF Conta Garantida'),
    (8528323193, 'Omie.CASH'),
    (8650909950, 'GETNET'),
    (8591664847, 'Itaú Unibanco'),
    (8591664978, 'CEF 2'),
]

CATEGORIAS_ALVO = {
    'mensalidade', 'pontuação', 'saldo de pontuação',
    'ipcd', 'distrato - aviso prévio',
}

hoje     = date.today()
d_inicio = hoje - timedelta(days=90)
PERIODO_INICIO = d_inicio.strftime('%d/%m/%Y')
PERIODO_FIM    = hoje.strftime('%d/%m/%Y')

def norm(s):
    if not s: return ''
    s = unicodedata.normalize('NFD', str(s))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower().strip()

TAGS_REGIAO = {
    'regiao abc', 'regiao alphaville', 'regiao litoral paulista',
    'regiao oeste e sul', 'regiao sorocaba', 'sao paulo',
    'zona leste', 'zona norte', 'guarulhos', 'litoral sul',
    'alagoas', 'belo horizonte', 'brasilia', 'curitiba',
    'rio de janeiro', 'porto alegre', 'fortaleza', 'recife',
    'salvador', 'manaus', 'belem', 'goiania', 'florianopolis',
    'campo grande', 'maceio', 'natal', 'teresina',
    'santa catarina', 'parana', 'noroeste paulista',
    'interior sp', 'grande sp', 'litoral', 'sul', 'minas gerais'
}

TAGS_MODELO = {
    'associados pro', 'associados', 'design',
    'cessao de direitos', 'associado pro', 'associado'
}

def extrair_caracteristicas(caracteristicas):
    dados = {'diretora_regional': '', 'regiao': '', 'status': '', 'modelo': '', 'gestora': ''}
    if not caracteristicas: return dados
    for c in caracteristicas:
        campo = norm(c.get('campo', ''))
        conteudo = str(c.get('conteudo', '')).strip()
        if any(x in campo for x in ['diretor', 'franquia', 'regional']):
            if 'regiao' not in campo:
                dados['diretora_regional'] = conteudo
        if 'gestor' in campo: dados['gestora'] = conteudo
        elif 'regiao' in campo: dados['regiao'] = conteudo
        elif 'status' in campo: dados['status'] = conteudo
        elif 'modelo' in campo: dados['modelo'] = conteudo
    return dados

def invalido(c):
    cnpj = c.get('cnpj_cpf', '') or ''
    nome = c.get('razao_social', '') or ''
    return (not cnpj or cnpj == '000.000.000-00'
            or not nome.strip() or nome == 'Cliente Consumidor')

def omie_post(url, body):
    for i in range(3):
        try:
            r = requests.post(url, json=body, timeout=30)
            return r.json()
        except Exception:
            if i == 2: raise
            time.sleep(2 ** i)

def calcular_dias(data_br):
    if not data_br: return 0
    try:
        d, m, y = data_br.split('/')
        venc = date(int(y), int(m), int(d))
        return max(0, (date.today() - venc).days)
    except:
        return 0

def extrair_ano(data_br):
    try:
        return int(data_br.split('/')[2])
    except:
        return 0

print('=' * 60)
print('CLUB&CASA - Gerando relatorio inadimplentes + recebimentos')
print('=' * 60)

# ----------------------------------------------------------------
# [1/5] LOJISTAS
# ----------------------------------------------------------------
print('\n[1/5] Buscando lojistas no Omie...')
p, tp = 1, 1
lojistas = {}
while p <= tp:
    r = omie_post(OMIE_CLIENTES, {
        'call': 'ListarClientes', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
        'param': [{'pagina': p, 'registros_por_pagina': RPP, 'apenas_importado_api': 'N', 'exibir_caracteristicas': 'S'}]
    })
    if not r or not r.get('clientes_cadastro'): break
    tp = r.get('total_de_paginas', 1)
    for c in r['clientes_cadastro']:
        if invalido(c): continue
        raw_tags = [t.get('tag', '') for t in c.get('tags', [])]
        nd_tags  = [norm(t) for t in raw_tags]
        if 'fornecedor' in nd_tags and 'cliente' not in nd_tags:
            continue
        caracs = extrair_caracteristicas(c.get('caracteristicas', []))
        distrato_tag = 'distrato' in nd_tags

        # Separa tags de região das demais
        tags_regiao_raw  = [raw_tags[i] for i, t in enumerate(nd_tags) if t in TAGS_REGIAO]
        tags_outras_raw  = [raw_tags[i] for i, t in enumerate(nd_tags) if t not in TAGS_REGIAO]

        regiao_tag   = tags_regiao_raw[0] if tags_regiao_raw else ''
        modelo_tag   = next((raw_tags[i] for i, t in enumerate(nd_tags) if t in TAGS_MODELO), '')

        regiao_final   = caracs['regiao'] if caracs['regiao'] else regiao_tag
        modelo_final   = caracs['modelo'] if caracs['modelo'] else modelo_tag
        distrato_final = 'Sim' if (norm(caracs['status']) == 'distrato' or distrato_tag) else 'Nao'

        email_raw = c.get('email', '') or ''
        codigo    = c.get('codigo_cliente_omie')

        lojistas[str(codigo)] = {
            'codigo':            codigo,
            'razao_social':      html_lib.unescape(c.get('razao_social', '') or ''),
            'nome_fantasia':     html_lib.unescape(c.get('nome_fantasia', '') or ''),
            'cnpj_cpf':          c.get('cnpj_cpf', '') or '',
            'email':             email_raw.split(',')[0].strip(),
            'telefone':          c.get('telefone1_numero', '') or '',
            'cidade':            c.get('cidade', '') or '',
            'estado':            c.get('estado', '') or '',
            'regiao':            regiao_final if regiao_final else 'Sem Regiao',
            'diretora_regional': caracs['diretora_regional'] if caracs['diretora_regional'] else 'Nao Informada',
            'gestora_carac':     caracs['gestora'] if caracs['gestora'] else '',
            'modelo':            modelo_final if modelo_final else 'Sem Modelo',
            'distrato':          distrato_final,
            'tags_regiao':       ', '.join(tags_regiao_raw),
            'tags_outras':       ', '.join(tags_outras_raw),
            'todas_tags':        ', '.join(raw_tags),
            'codigo_vendedor':   str((c.get('recomendacoes') or {}).get('codigo_vendedor', '') or ''),
        }
    print(f'  Pag {p:3}/{tp} - {len(lojistas)} lojistas')
    p += 1
    time.sleep(0.3)

# ----------------------------------------------------------------
# [2/5] GESTORAS
# ----------------------------------------------------------------
print('\n[2/5] Buscando gestoras...')
vendedores = {}
try:
    p, tp = 1, 1
    while p <= tp:
        r = omie_post(OMIE_VENDEDORES, {
            'call': 'ListarVendedores', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
            'param': [{'pagina': p, 'registros_por_pagina': 50}]
        })
        if not r or not r.get('cadastro'): break
        tp = r.get('total_de_paginas', 1)
        for v in r['cadastro']:
            cod  = str(v.get('codigo', ''))
            nome = v.get('nome', '') or ''
            if cod and nome: vendedores[cod] = nome
        p += 1
        time.sleep(0.3)
    print(f'  {len(vendedores)} gestoras encontradas')
except Exception as e:
    print(f'  Erro: {e}')

# ----------------------------------------------------------------
# [2b/5] CATEGORIAS
# ----------------------------------------------------------------
print('\n[2b/5] Buscando categorias...')
categorias_map = {}
try:
    p, tp = 1, 1
    while p <= tp:
        r = omie_post(OMIE_CATEGORIAS, {
            'call': 'ListarCategorias', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
            'param': [{'pagina': p, 'registros_por_pagina': 100}]
        })
        if not r or not r.get('categoria_cadastro'): break
        tp = r.get('total_de_paginas', 1)
        for cat in r['categoria_cadastro']:
            cod  = cat.get('codigo', '') or cat.get('codigo_categoria', '')
            desc = cat.get('descricao', '') or ''
            if cod and desc: categorias_map[cod] = desc
        p += 1
        time.sleep(0.3)
    print(f'  {len(categorias_map)} categorias encontradas')
except Exception as e:
    print(f'  Erro: {e}')

# ----------------------------------------------------------------
# [3/5] TÍTULOS ATRASADOS
# ----------------------------------------------------------------
print('\n[3/5] Buscando titulos ATRASADOS...')
p, tp = 1, 1
titulos = {}
total_tit = 0
total_cancelados = 0
erros_consecutivos = 0

while p <= tp:
    r = omie_post(OMIE_CONTAS, {
        'call': 'ListarContasReceber', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
        'param': [{'pagina': p, 'registros_por_pagina': RPP, 'filtrar_por_status': 'ATRASADO'}]
    })
    if not r:
        print(f'  Sem resposta na pag {p}, encerrando.')
        break
    if r.get('faultstring'):
        erros_consecutivos += 1
        print(f'  Aviso pag {p}: {r["faultstring"]}')
        if erros_consecutivos >= 5: break
        p += 1; time.sleep(1.5); continue
    erros_consecutivos = 0
    tp   = r.get('total_de_paginas', 1)
    regs = r.get('conta_receber_cadastro') or r.get('lista_contareceber') or []
    for t in regs:
        cod = str(t.get('codigo_cliente_fornecedor', ''))
        if not cod: continue
        status_tit = (t.get('status_titulo') or '').upper().strip()
        if status_tit == 'CANCELADO':
            total_cancelados += 1
            continue
        if cod not in titulos: titulos[cod] = []
        titulos[cod].append(t)
        total_tit += 1
    print(f'  Pag {p:3}/{tp} - {total_tit} titulos | {len(titulos)} clientes')
    p += 1
    time.sleep(0.6)

# ----------------------------------------------------------------
# [4/5] EXTRATO - RECEBIMENTOS (últimos 90 dias)
# ----------------------------------------------------------------
print(f'\n[4/5] Buscando recebimentos ({PERIODO_INICIO} a {PERIODO_FIM})...')
recebimentos = []

for nCodCC, nome_conta in CONTAS_RECEBIMENTO:
    try:
        r = omie_post(OMIE_EXTRATO, {
            'call': 'ListarExtrato', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
            'param': [{'nCodCC': nCodCC, 'dPeriodoInicial': PERIODO_INICIO, 'dPeriodoFinal': PERIODO_FIM}]
        })
        movs = r.get('listaMovimentos', [])
        cnt = 0
        for m in movs:
            if m.get('cNatureza') != 'R': continue
            cat = m.get('cDesCategoria', '').lower().strip()
            if cat not in CATEGORIAS_ALVO: continue
            valor = float(m.get('nValorDocumento', 0))
            if valor <= 0: continue
            recebimentos.append({
                'Data':           m.get('dDataLancamento', ''),
                'Conciliacao':    m.get('dDataConciliacao', ''),
                'Cliente':        m.get('cDesCliente', '') or m.get('cRazCliente', ''),
                'CNPJ':           m.get('cDocCliente', ''),
                'Categoria':      m.get('cDesCategoria', ''),
                'Conta Corrente': nome_conta,
                'Documento':      m.get('cNumero', '') or m.get('cDocumentoFiscal', ''),
                'Tipo Doc':       m.get('cTipoDocumento', ''),
                'Situacao':       m.get('cSituacao', ''),
                'Vendedor':       m.get('cVendedor', ''),
                'Valor':          valor,
                'Observacoes':    m.get('cObservacoes', ''),
            })
            cnt += 1
        print(f'  {nome_conta}: {cnt} recebimentos')
    except Exception as e:
        print(f'  {nome_conta}: erro - {e}')
    time.sleep(0.5)

recebimentos.sort(key=lambda x: x['Data'], reverse=True)
total_recebido = round(sum(r['Valor'] for r in recebimentos), 2)
print(f'  Total: {len(recebimentos)} registros | R$ {total_recebido:,.2f}')

# ----------------------------------------------------------------
# [5/5] CRUZAMENTO INADIMPLENTES
# ----------------------------------------------------------------
print('\n[5/5] Cruzando dados...')

def sort_key(t):
    dv = t.get('data_vencimento') or '99/99/9999'
    try:
        d, m, y = dv.split('/')
        return (int(y), int(m), int(d))
    except:
        return (9999, 99, 99)

status_map = {'ATRASADO': 'Atrasado', 'PAGO': 'Pago', 'ABERTO': 'A Vencer', 'CANCELADO': 'Cancelado'}

rows_xlsx = []
rows_json = []
total_valor    = 0
total_distrato = 0

for cod, tits in titulos.items():
    cl = lojistas.get(cod)
    if not cl: continue

    tits_sorted = sorted(tits, key=sort_key)
    val_total   = round(sum(t.get('valor_documento', 0) for t in tits), 2)
    mais_antigo = tits_sorted[0].get('data_vencimento', '')

    gestora_nome = cl['gestora_carac']
    if not gestora_nome and cl['codigo_vendedor']:
        gestora_nome = vendedores.get(cl['codigo_vendedor'], cl['codigo_vendedor'])

    total_valor += val_total
    if cl['distrato'] == 'Sim':
        total_distrato += val_total

    titulos_arr = []
    for t in tits_sorted:
        data_venc = t.get('data_vencimento') or ''
        valor_tit = float(t.get('valor_documento') or t.get('nValParcela') or 0)
        numero_doc = (
            t.get('numero_documento_fiscal') or
            t.get('numero_documento') or
            t.get('cNumParcela') or
            str(t.get('nIdTitulo', '')) or 'N/I'
        )
        cod_categoria = t.get('codigo_categoria') or ''
        if not cod_categoria:
            categorias_arr = t.get('categorias') or []
            if categorias_arr:
                cod_categoria = categorias_arr[0].get('codigo_categoria', '')
        categoria   = categorias_map.get(cod_categoria, cod_categoria)
        status_raw  = (t.get('status_titulo') or t.get('cStatus') or 'ATRASADO').upper().strip()
        status_exib = status_map.get(status_raw, status_raw.capitalize())
        ano_venc    = extrair_ano(data_venc)

        titulos_arr.append({
            'numero_documento': numero_doc,
            'categoria':        categoria,
            'data_vencimento':  data_venc,
            'ano':              ano_venc,
            'status':           status_exib,
            'dias_atraso':      calcular_dias(data_venc),
            'valor_documento':  valor_tit,
        })

        # Excel: 1 linha por título
        rows_xlsx.append({
            'Razao Social':           cl['razao_social'],
            'Nome Fantasia':          cl['nome_fantasia'],
            'CNPJ/CPF':               cl['cnpj_cpf'],
            'Email':                  cl['email'],
            'Telefone':               cl['telefone'],
            'Cidade':                 cl['cidade'],
            'Estado':                 cl['estado'],
            'Regiao':                 cl['regiao'],
            'Tags Regiao':            cl['tags_regiao'],
            'Outras Tags':            cl['tags_outras'],
            'Diretora Regional':      cl['diretora_regional'],
            'Gestora':                gestora_nome or '',
            'Modelo de Negocio':      cl['modelo'],
            'Distrato':               cl['distrato'],
            'Qtd Titulos':            len(tits),
            'Valor Atrasado (R$)':    val_total,
            'Dias em Atraso':         calcular_dias(mais_antigo),
            'Vencimento Mais Antigo': mais_antigo,
            'Ano Vencimento':         ano_venc,
            'Status':                 status_exib,
            'Documento':              numero_doc,
            'Categoria':              categoria,
            'Vencimento':             data_venc,
            'Valor Titulo (R$)':      valor_tit,
            'Codigo Omie':            cl['codigo'],
        })

    # JSON: 1 linha por cliente
    rows_json.append({
        'Razao Social':           cl['razao_social'],
        'Nome Fantasia':          cl['nome_fantasia'],
        'CNPJ/CPF':               cl['cnpj_cpf'],
        'Email':                  cl['email'],
        'Telefone':               cl['telefone'],
        'Cidade':                 cl['cidade'],
        'Estado':                 cl['estado'],
        'Regiao':                 cl['regiao'],
        'Tags Regiao':            cl['tags_regiao'],
        'Outras Tags':            cl['tags_outras'],
        'Diretora Regional':      cl['diretora_regional'],
        'Gestora':                gestora_nome or '',
        'Modelo de Negocio':      cl['modelo'],
        'Distrato':               cl['distrato'],
        'Qtd Titulos':            len(tits),
        'Valor Atrasado (R$)':    val_total,
        'Dias em Atraso':         calcular_dias(mais_antigo),
        'Vencimento Mais Antigo': mais_antigo,
        'Codigo Omie':            cl['codigo'],
        'titulos':                titulos_arr,
    })

rows_xlsx.sort(key=lambda x: (-x['Valor Atrasado (R$)'], x['Razao Social'], x['Vencimento']))
rows_json.sort(key=lambda x: -x['Valor Atrasado (R$)'])
total_valor    = round(total_valor, 2)
total_distrato = round(total_distrato, 2)

# ----------------------------------------------------------------
# SALVAR JSONs
# ----------------------------------------------------------------
print('\nSalvando inadimplentes.json...')
with open('inadimplentes.json', 'w', encoding='utf-8') as f:
    json.dump({
        'gerado_em':      datetime.now().strftime('%d/%m/%Y %H:%M'),
        'total_clientes': len(rows_json),
        'total_titulos':  total_tit,
        'valor_total':    total_valor,
        'inadimplentes':  rows_json,
    }, f, ensure_ascii=False, indent=2)

print('Salvando recebimentos.json...')
with open('recebimentos.json', 'w', encoding='utf-8') as f:
    json.dump({
        'gerado_em':       datetime.now().strftime('%d/%m/%Y %H:%M'),
        'periodo_inicio':  PERIODO_INICIO,
        'periodo_fim':     PERIODO_FIM,
        'total_registros': len(recebimentos),
        'valor_total':     total_recebido,
        'recebimentos':    recebimentos,
    }, f, ensure_ascii=False, indent=2)

# ----------------------------------------------------------------
# SALVAR XLSX com 2 abas
# ----------------------------------------------------------------
print('Salvando inadimplentes-clubcasa.xlsx...')

wb = Workbook()
roxo        = PatternFill(start_color='5B52E8', end_color='5B52E8', fill_type='solid')
roxo_claro  = PatternFill(start_color='F4F6FB', end_color='F4F6FB', fill_type='solid')
verde       = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
verde_claro = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
branco_ft   = Font(color='FFFFFF', bold=True, size=11, name='Arial')
borda = Border(
    left=Side(style='thin', color='E2E6F0'), right=Side(style='thin', color='E2E6F0'),
    top=Side(style='thin', color='E2E6F0'),  bottom=Side(style='thin', color='E2E6F0')
)

# ── ABA 1: Inadimplentes (título a título) ───────────────────────
ws1 = wb.active
ws1.title = 'Inadimplentes'

headers1 = [
    'Razao Social', 'Nome Fantasia', 'CNPJ/CPF', 'Email', 'Telefone',
    'Cidade', 'Estado', 'Regiao', 'Tags Regiao', 'Outras Tags',
    'Diretora Regional', 'Gestora', 'Modelo de Negocio', 'Distrato',
    'Qtd Titulos', 'Valor Atrasado (R$)', 'Dias em Atraso', 'Vencimento Mais Antigo',
    'Ano Vencimento', 'Status', 'Documento', 'Categoria', 'Vencimento',
    'Valor Titulo (R$)', 'Codigo Omie'
]

for col, h in enumerate(headers1, start=1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.fill = roxo; cell.font = branco_ft
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = borda
ws1.row_dimensions[1].height = 30

prev_cnpj = None; usar_claro = False
for i, r in enumerate(rows_xlsx, start=2):
    if r['CNPJ/CPF'] != prev_cnpj:
        usar_claro = not usar_claro; prev_cnpj = r['CNPJ/CPF']
    valores = [
        r['Razao Social'], r['Nome Fantasia'], r['CNPJ/CPF'], r['Email'], r['Telefone'],
        r['Cidade'], r['Estado'], r['Regiao'], r['Tags Regiao'], r['Outras Tags'],
        r['Diretora Regional'], r['Gestora'], r['Modelo de Negocio'], r['Distrato'],
        r['Qtd Titulos'], r['Valor Atrasado (R$)'], r['Dias em Atraso'], r['Vencimento Mais Antigo'],
        r['Ano Vencimento'], r['Status'], r['Documento'], r['Categoria'], r['Vencimento'],
        r['Valor Titulo (R$)'], r['Codigo Omie']
    ]
    for col, v in enumerate(valores, start=1):
        cell = ws1.cell(row=i, column=col, value=v)
        cell.border = borda; cell.font = Font(name='Arial', size=10)
        if usar_claro: cell.fill = roxo_claro
        if col == 16: cell.number_format = 'R$ #,##0.00'
        if col == 24: cell.number_format = 'R$ #,##0.00'

ultima1 = len(rows_xlsx) + 2
for col in range(1, len(headers1) + 1):
    cell = ws1.cell(row=ultima1, column=col)
    cell.fill = roxo; cell.border = borda; cell.font = branco_ft
ws1.cell(row=ultima1, column=1, value='TOTAL GERAL')
c = ws1.cell(row=ultima1, column=24, value=round(sum(r['Valor Titulo (R$)'] for r in rows_xlsx), 2))
c.number_format = 'R$ #,##0.00'; c.fill = roxo; c.font = branco_ft

larguras1 = [35,30,18,30,15,20,8,22,22,30,20,22,20,10,10,18,12,18,8,12,20,22,14,16,14]
for i, w in enumerate(larguras1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f'A1:{get_column_letter(len(headers1))}1'

# ── ABA 2: Recebimentos ──────────────────────────────────────────
ws2 = wb.create_sheet('Recebimentos')

headers2 = [
    'Data Lançamento', 'Data Conciliação', 'Cliente', 'CNPJ',
    'Categoria', 'Conta Corrente', 'Documento', 'Tipo Doc',
    'Situação', 'Vendedor/Gestora', 'Valor (R$)', 'Observações'
]

for col, h in enumerate(headers2, start=1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = verde; cell.font = branco_ft
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = borda
ws2.row_dimensions[1].height = 30

for i, r in enumerate(recebimentos, start=2):
    valores = [
        r['Data'], r['Conciliacao'], r['Cliente'], r['CNPJ'],
        r['Categoria'], r['Conta Corrente'], r['Documento'], r['Tipo Doc'],
        r['Situacao'], r['Vendedor'], r['Valor'], r['Observacoes']
    ]
    for col, v in enumerate(valores, start=1):
        cell = ws2.cell(row=i, column=col, value=v)
        cell.border = borda; cell.font = Font(name='Arial', size=10)
        if i % 2 == 0: cell.fill = verde_claro
        if col == 11: cell.number_format = 'R$ #,##0.00'

ultima2 = len(recebimentos) + 2
for col in range(1, len(headers2) + 1):
    cell = ws2.cell(row=ultima2, column=col)
    cell.fill = verde; cell.border = borda; cell.font = branco_ft
ws2.cell(row=ultima2, column=1, value='TOTAL RECEBIDO')
c = ws2.cell(row=ultima2, column=11, value=total_recebido)
c.number_format = 'R$ #,##0.00'; c.fill = verde; c.font = branco_ft

larguras2 = [14,14,35,18,20,20,18,12,16,20,14,40]
for i, w in enumerate(larguras2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(headers2))}1'

wb.save('inadimplentes-clubcasa.xlsx')

print(f'\n{"=" * 60}')
print(f'CONCLUIDO!')
print(f'Clientes inadimplentes: {len(rows_json)}')
print(f'Total de titulos:       {total_tit}')
print(f'Valor total atrasado:   R$ {total_valor:,.2f}')
print(f'Valor em distrato:      R$ {total_distrato:,.2f}')
print(f'Recebimentos (90d):     {len(recebimentos)} | R$ {total_recebido:,.2f}')
print(f'Titulos cancelados:     {total_cancelados}')
print(f'Arquivos: inadimplentes.json | recebimentos.json | inadimplentes-clubcasa.xlsx')