"""
rodar.py — Club&Casa Design
Roda todo dia automaticamente:
1. Gera inadimplentes.json + recebimentos.json
2. Atualiza Google Sheets (4 abas: Inadimplentes, Recebidos, A Receber, Dashboard)
3. Publica no Vercel (git push)

Agendar no Windows: Task Scheduler → rodar.py todo dia às 7h
"""
import requests, time, json, html as html_lib, unicodedata, subprocess, os
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import gspread
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pickle

# ─── CONFIGURAÇÕES ────────────────────────────────────────────────
APP_KEY    = '1576318757013'
APP_SECRET = '2d96d060ad7abe5dd82c4d214f3a9de8'
SHEET_ID   = '1FzzxksfyE0K7PHMGGuLxrAFL35mTtag80Avr97B1Cjw'
CLIENT_FILE = 'client_secret.json'
TOKEN_FILE  = 'token_sheets.pkl'
SCOPES      = ['https://www.googleapis.com/auth/spreadsheets']

OMIE_CLIENTES   = 'https://app.omie.com.br/api/v1/geral/clientes/'
OMIE_CONTAS     = 'https://app.omie.com.br/api/v1/financas/contareceber/'
OMIE_VENDEDORES = 'https://app.omie.com.br/api/v1/geral/vendedores/'
OMIE_CATEGORIAS = 'https://app.omie.com.br/api/v1/geral/categorias/'
OMIE_EXTRATO    = 'https://app.omie.com.br/api/v1/financas/extrato/'
RPP = 100

# Todas as contas correntes com movimento relevante
CONTAS_RECEBIMENTO = [
    (2334749009, 'Caixinha'),
    (2337038431, 'Bradesco'),
    (2337038587, 'Santander Conta Corrente'),
    (2337042886, 'Caixa Econômica Federal'),
    (2526653437, 'Perda'),
    (8116675375, 'Itaú Unibanco Desconto'),
    (8445110562, 'ASAAS'),
    (8450846510, 'Permuta'),
    (8453130413, 'CEF Conta Garantida'),
    (8528323193, 'Omie.CASH'),
    (8541907465, 'EXTRAJUDICIAL INATIVOS'),
    (8549373253, 'Boletos Fraudados'),
    (8591664847, 'Itaú Unibanco2'),
    (8650909950, 'GETNET'),
]

CATEGORIAS_RECEB = {
    'mensalidade', 'pontuação', 'saldo de pontuação',
    'ipcd', 'distrato - aviso prévio', 'acordo', 'reacordo'
}

TAGS_REGIAO = {
    'regiao abc','regiao alphaville','regiao litoral paulista','regiao oeste e sul',
    'regiao sorocaba','sao paulo','zona leste','zona norte','guarulhos','litoral sul',
    'alagoas','belo horizonte','brasilia','curitiba','rio de janeiro','porto alegre',
    'fortaleza','recife','salvador','manaus','belem','goiania','florianopolis',
    'campo grande','maceio','natal','teresina','santa catarina','parana',
    'noroeste paulista','interior sp','grande sp','litoral','sul','minas gerais'
}
TAGS_MODELO = {
    'associados pro','associados','design','cessao de direitos','associado pro','associado'
}

hoje      = date.today()
hoje_str  = hoje.strftime('%d/%m/%Y')
ano_atual = hoje.year

print('='*60)
print(f'CLUB&CASA — Automação Diária — {hoje_str}')
print('='*60)

# ─── HELPERS ──────────────────────────────────────────────────────
def norm(s):
    if not s: return ''
    s = unicodedata.normalize('NFD', str(s))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower().strip()

def omie_post(url, call, param):
    body = {'call': call, 'app_key': APP_KEY, 'app_secret': APP_SECRET, 'param': [param]}
    for i in range(3):
        try:
            r = requests.post(url, json=body, timeout=30)
            return r.json()
        except:
            if i == 2: raise
            time.sleep(2**i)

def calcular_dias(data_br):
    if not data_br: return 0
    try:
        d, m, y = data_br.split('/')
        return max(0, (date.today() - date(int(y), int(m), int(d))).days)
    except: return 0

def extrair_ano(data_br):
    try: return int(data_br.split('/')[2])
    except: return 0

def unescape(s):
    if not s: return ''
    return html_lib.unescape(str(s))

def extrair_caracteristicas(caracs):
    dados = {'diretora_regional':'','regiao':'','status':'','modelo':'','gestora':''}
    for c in (caracs or []):
        campo = norm(c.get('campo',''))
        conteudo = str(c.get('conteudo','')).strip()
        if any(x in campo for x in ['diretor','franquia','regional']):
            if 'regiao' not in campo: dados['diretora_regional'] = conteudo
        if 'gestor' in campo: dados['gestora'] = conteudo
        elif 'regiao' in campo: dados['regiao'] = conteudo
        elif 'status' in campo: dados['status'] = conteudo
        elif 'modelo' in campo: dados['modelo'] = conteudo
    return dados

def invalido(c):
    cnpj = c.get('cnpj_cpf','') or ''
    nome = c.get('razao_social','') or ''
    return not cnpj or cnpj == '000.000.000-00' or not nome.strip() or nome == 'Cliente Consumidor'

# ─── GOOGLE SHEETS AUTH ───────────────────────────────────────────
def get_gc():
    creds = None
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE,'rb') as f: creds = pickle.load(f)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE,'wb') as f: pickle.dump(creds, f)
    return gspread.authorize(creds)

def garantir_aba(spreadsheet, nome, cabecalho):
    try:
        ws = spreadsheet.worksheet(nome)
    except:
        ws = spreadsheet.add_worksheet(title=nome, rows=10000, cols=len(cabecalho)+2)
    ws.clear()
    ws.append_row(cabecalho)
    return ws

# ─── [1/5] LOJISTAS ───────────────────────────────────────────────
print('\n[1/5] Buscando lojistas...')
p, tp = 1, 1
lojistas = {}
while p <= tp:
    r = omie_post(OMIE_CLIENTES, 'ListarClientes', {
        'pagina': p, 'registros_por_pagina': RPP,
        'apenas_importado_api': 'N', 'exibir_caracteristicas': 'S'
    })
    if not r or not r.get('clientes_cadastro'): break
    tp = r.get('total_de_paginas', 1)
    for c in r['clientes_cadastro']:
        if invalido(c): continue
        raw_tags = [t.get('tag','') for t in c.get('tags',[])]
        nd_tags  = [norm(t) for t in raw_tags]
        if 'fornecedor' in nd_tags and 'cliente' not in nd_tags: continue
        caracs = extrair_caracteristicas(c.get('caracteristicas',[]))
        tags_regiao = [raw_tags[i] for i,t in enumerate(nd_tags) if t in TAGS_REGIAO]
        tags_outras = [raw_tags[i] for i,t in enumerate(nd_tags) if t not in TAGS_REGIAO]
        regiao_tag  = tags_regiao[0] if tags_regiao else ''
        modelo_tag  = next((raw_tags[i] for i,t in enumerate(nd_tags) if t in TAGS_MODELO), '')
        distrato_tag = 'distrato' in nd_tags
        regiao_final   = caracs['regiao'] or regiao_tag or 'Sem Regiao'
        modelo_final   = caracs['modelo'] or modelo_tag or 'Sem Modelo'
        distrato_final = 'Sim' if (norm(caracs['status']) == 'distrato' or distrato_tag) else 'Nao'
        email_raw = c.get('email','') or ''
        # Filtra email do Club Casa
        emails = [e.strip() for e in email_raw.split(',')]
        email  = next((e for e in emails if e and 'clubecasadesign' not in e.lower()), '')
        codigo = c.get('codigo_cliente_omie')
        lojistas[str(codigo)] = {
            'codigo': codigo,
            'razao_social': unescape(c.get('razao_social','')),
            'nome_fantasia': unescape(c.get('nome_fantasia','')),
            'cnpj_cpf': c.get('cnpj_cpf',''),
            'email': email,
            'telefone': c.get('telefone1_numero',''),
            'telefone_ddd': c.get('telefone1_ddd',''),
            'cidade': c.get('cidade',''),
            'estado': c.get('estado',''),
            'regiao': regiao_final,
            'tags_regiao': ', '.join(tags_regiao),
            'tags_outras': ', '.join(tags_outras),
            'diretora_regional': caracs['diretora_regional'] or 'Nao Informada',
            'gestora_carac': caracs['gestora'],
            'modelo': modelo_final,
            'distrato': distrato_final,
            'todas_tags': ', '.join(raw_tags),
            'codigo_vendedor': str((c.get('recomendacoes') or {}).get('codigo_vendedor','') or ''),
        }
    print(f'  Pag {p}/{tp} — {len(lojistas)} lojistas')
    p += 1
    time.sleep(0.3)

# ─── [2/5] GESTORAS ───────────────────────────────────────────────
print('\n[2/5] Buscando gestoras...')
vendedores = {}
p, tp = 1, 1
while p <= tp:
    r = omie_post(OMIE_VENDEDORES, 'ListarVendedores', {'pagina': p, 'registros_por_pagina': 50})
    if not r or not r.get('cadastro'): break
    tp = r.get('total_de_paginas', 1)
    for v in r['cadastro']:
        if v.get('codigo') and v.get('nome'):
            vendedores[str(v['codigo'])] = v['nome']
    p += 1
    time.sleep(0.2)
print(f'  {len(vendedores)} gestoras')

# ─── [2b] CATEGORIAS ──────────────────────────────────────────────
print('\n[2b] Buscando categorias...')
categorias_map = {}
p, tp = 1, 1
while p <= tp:
    r = omie_post(OMIE_CATEGORIAS, 'ListarCategorias', {'pagina': p, 'registros_por_pagina': 100})
    if not r or not r.get('categoria_cadastro'): break
    tp = r.get('total_de_paginas', 1)
    for cat in r['categoria_cadastro']:
        cod = cat.get('codigo','')
        desc = cat.get('descricao','')
        if cod and desc: categorias_map[cod] = desc
    p += 1
    time.sleep(0.2)
print(f'  {len(categorias_map)} categorias')

# ─── [3/5] TÍTULOS ATRASADOS ──────────────────────────────────────
print('\n[3/5] Buscando títulos ATRASADOS...')
titulos_inad = {}
total_tit = 0
p, tp = 1, 1
erros = 0
while p <= tp:
    r = omie_post(OMIE_CONTAS, 'ListarContasReceber', {
        'pagina': p, 'registros_por_pagina': RPP, 'filtrar_por_status': 'ATRASADO'
    })
    if not r:
        print(f'  Sem resposta pag {p}'); break
    if r.get('faultstring'):
        erros += 1
        if erros >= 5: break
        p += 1; time.sleep(1.5); continue
    erros = 0
    tp = r.get('total_de_paginas', 1)
    for t in (r.get('conta_receber_cadastro') or []):
        cod = str(t.get('codigo_cliente_fornecedor',''))
        if not cod or t.get('status_titulo','').upper() == 'CANCELADO': continue
        if cod not in titulos_inad: titulos_inad[cod] = []
        titulos_inad[cod].append(t)
        total_tit += 1
    print(f'  Pag {p}/{tp} — {total_tit} títulos')
    p += 1
    time.sleep(0.5)

# ─── [4/5] A RECEBER (abertos futuros) ───────────────────────────
print('\n[4/5] Buscando A RECEBER (abertos)...')
titulos_avencer = {}
total_avencer = 0
p, tp = 1, 1
erros = 0
inicio_avencer = hoje_str
fim_avencer    = f'31/12/{ano_atual}'
while p <= tp:
    r = omie_post(OMIE_CONTAS, 'ListarContasReceber', {
        'pagina': p, 'registros_por_pagina': RPP,
        'filtrar_apenas_titulos_em_aberto': 'S',
        'data_de': inicio_avencer,
        'data_ate': fim_avencer,
    })
    if not r:
        break
    if r.get('faultstring'):
        erros += 1
        if erros >= 5: break
        p += 1; time.sleep(1.5); continue
    erros = 0
    tp = r.get('total_de_paginas', 1)
    for t in (r.get('conta_receber_cadastro') or []):
        cod = str(t.get('codigo_cliente_fornecedor',''))
        if not cod: continue
        dv = t.get('data_vencimento','')
        if calcular_dias(dv) > 0: continue  # já vencido = inadimplente
        if cod not in titulos_avencer: titulos_avencer[cod] = []
        titulos_avencer[cod].append(t)
        total_avencer += 1
    print(f'  Pag {p}/{tp} — {total_avencer} títulos a vencer')
    p += 1
    time.sleep(0.5)

# ─── [5/5] RECEBIMENTOS (extrato todas as contas) ────────────────
print('\n[5/5] Buscando recebimentos 2026 (todas as contas)...')
recebimentos = []
for nCodCC, nome_conta in CONTAS_RECEBIMENTO:
    try:
        r = omie_post(OMIE_EXTRATO, 'ListarExtrato', {
            'nCodCC': nCodCC,
            'dPeriodoInicial': f'01/01/{ano_atual}',
            'dPeriodoFinal': hoje_str
        })
        movs = r.get('listaMovimentos', []) if r else []
        cnt = 0
        for m in movs:
            if m.get('cNatureza') != 'R': continue
            cat = m.get('cDesCategoria','').lower().strip()
            if cat not in CATEGORIAS_RECEB: continue
            valor = float(m.get('nValorDocumento', 0))
            if valor <= 0: continue
            recebimentos.append({
                'Data': m.get('dDataLancamento',''),
                'Conciliacao': m.get('dDataConciliacao',''),
                'Cliente': m.get('cDesCliente','') or m.get('cRazCliente',''),
                'CNPJ': m.get('cDocCliente',''),
                'Categoria': m.get('cDesCategoria',''),
                'Conta Corrente': nome_conta,
                'Documento': m.get('cNumero','') or m.get('cDocumentoFiscal',''),
                'Tipo Doc': m.get('cTipoDocumento',''),
                'Situacao': m.get('cSituacao',''),
                'Vendedor': m.get('cVendedor',''),
                'Valor': valor,
            })
            cnt += 1
        print(f'  {nome_conta}: {cnt} registros')
    except Exception as e:
        print(f'  {nome_conta}: erro — {e}')
    time.sleep(0.4)
recebimentos.sort(key=lambda x: x['Data'], reverse=True)
total_recebido = sum(r['Valor'] for r in recebimentos)
print(f'  TOTAL RECEBIDO {ano_atual}: {len(recebimentos)} lançamentos | R$ {total_recebido:,.2f}')

# ─── CRUZAMENTO INADIMPLENTES ─────────────────────────────────────
print('\nCruzando inadimplentes...')

def sort_key(t):
    dv = t.get('data_vencimento','') or '99/99/9999'
    try:
        d,m,y = dv.split('/')
        return (int(y),int(m),int(d))
    except: return (9999,99,99)

status_map = {'ATRASADO':'Atrasado','PAGO':'Pago','ABERTO':'A Vencer','CANCELADO':'Cancelado'}

rows_inad_json = []
rows_inad_xlsx = []
total_valor = 0
total_distrato = 0

for cod, tits in titulos_inad.items():
    cl = lojistas.get(cod)
    if not cl: continue
    tits_sorted = sorted(tits, key=sort_key)
    val_total   = round(sum(t.get('valor_documento',0) for t in tits), 2)
    mais_antigo = tits_sorted[0].get('data_vencimento','')
    gestora = cl['gestora_carac'] or vendedores.get(cl['codigo_vendedor'],'')
    total_valor += val_total
    if cl['distrato'] == 'Sim': total_distrato += val_total

    titulos_arr = []
    for t in tits_sorted:
        dv = t.get('data_vencimento','')
        vt = float(t.get('valor_documento',0))
        ndoc = (t.get('numero_documento_fiscal') or t.get('numero_documento') or
                t.get('cNumParcela') or str(t.get('nIdTitulo','')) or 'N/I')
        cod_cat = t.get('codigo_categoria','')
        categoria = categorias_map.get(cod_cat, cod_cat)
        st = (t.get('status_titulo') or 'ATRASADO').upper().strip()
        titulos_arr.append({
            'numero_documento': ndoc, 'categoria': categoria,
            'data_vencimento': dv, 'ano': extrair_ano(dv),
            'status': status_map.get(st, st.capitalize()),
            'dias_atraso': calcular_dias(dv), 'valor_documento': vt,
        })
        rows_inad_xlsx.append({
            'Razao Social': cl['razao_social'], 'Nome Fantasia': cl['nome_fantasia'],
            'CNPJ/CPF': cl['cnpj_cpf'], 'Email': cl['email'],
            'Telefone': cl['telefone'], 'Cidade': cl['cidade'], 'Estado': cl['estado'],
            'Regiao': cl['regiao'], 'Tags Regiao': cl['tags_regiao'],
            'Outras Tags': cl['tags_outras'], 'Diretora Regional': cl['diretora_regional'],
            'Gestora': gestora, 'Modelo de Negocio': cl['modelo'], 'Distrato': cl['distrato'],
            'Qtd Titulos': len(tits), 'Valor Atrasado (R$)': val_total,
            'Dias em Atraso': calcular_dias(mais_antigo), 'Vencimento Mais Antigo': mais_antigo,
            'Ano Vencimento': extrair_ano(dv), 'Status': status_map.get(st, st.capitalize()),
            'Documento': ndoc, 'Categoria': categoria, 'Vencimento': dv,
            'Valor Titulo (R$)': vt, 'Codigo Omie': cl['codigo'],
        })

    rows_inad_json.append({
        'Razao Social': cl['razao_social'], 'Nome Fantasia': cl['nome_fantasia'],
        'CNPJ/CPF': cl['cnpj_cpf'], 'Email': cl['email'],
        'Telefone': cl['telefone'], 'Cidade': cl['cidade'], 'Estado': cl['estado'],
        'Regiao': cl['regiao'], 'Tags Regiao': cl['tags_regiao'],
        'Outras Tags': cl['tags_outras'], 'Diretora Regional': cl['diretora_regional'],
        'Gestora': gestora, 'Modelo de Negocio': cl['modelo'], 'Distrato': cl['distrato'],
        'Qtd Titulos': len(tits), 'Valor Atrasado (R$)': val_total,
        'Dias em Atraso': calcular_dias(mais_antigo), 'Vencimento Mais Antigo': mais_antigo,
        'Codigo Omie': cl['codigo'], 'titulos': titulos_arr,
    })

rows_inad_json.sort(key=lambda x: -x['Valor Atrasado (R$)'])
rows_inad_xlsx.sort(key=lambda x: (-x['Valor Atrasado (R$)'], x['Razao Social'], x['Vencimento']))
total_valor    = round(total_valor, 2)
total_distrato = round(total_distrato, 2)

# ─── CRUZAMENTO A RECEBER ─────────────────────────────────────────
print('Cruzando a receber...')
rows_avencer = []
for cod, tits in titulos_avencer.items():
    cl = lojistas.get(cod)
    if not cl: continue
    gestora = cl['gestora_carac'] or vendedores.get(cl['codigo_vendedor'],'')
    for t in sorted(tits, key=sort_key):
        dv = t.get('data_vencimento','')
        vt = float(t.get('valor_documento',0))
        ndoc = (t.get('numero_documento_fiscal') or t.get('numero_documento') or 'N/I')
        cod_cat = t.get('codigo_categoria','')
        rows_avencer.append({
            'Razao Social': cl['razao_social'], 'Nome Fantasia': cl['nome_fantasia'],
            'CNPJ/CPF': cl['cnpj_cpf'], 'Email': cl['email'],
            'Cidade': cl['cidade'], 'Estado': cl['estado'],
            'Regiao': cl['regiao'], 'Gestora': gestora,
            'Modelo de Negocio': cl['modelo'], 'Distrato': cl['distrato'],
            'Documento': ndoc, 'Categoria': categorias_map.get(cod_cat, cod_cat),
            'Vencimento': dv, 'Ano': extrair_ano(dv),
            'Valor (R$)': vt, 'Codigo Omie': cl['codigo'],
        })
rows_avencer.sort(key=lambda x: x['Vencimento'])
total_avencer_valor = round(sum(r['Valor (R$)'] for r in rows_avencer), 2)

# ─── SALVAR JSONs ─────────────────────────────────────────────────
print('\nSalvando JSONs...')
gerado_em = datetime.now().strftime('%d/%m/%Y %H:%M')

with open('inadimplentes.json','w',encoding='utf-8') as f:
    json.dump({
        'gerado_em': gerado_em,
        'total_clientes': len(rows_inad_json),
        'total_titulos': total_tit,
        'valor_total': total_valor,
        'inadimplentes': rows_inad_json,
    }, f, ensure_ascii=False, indent=2)

with open('recebimentos.json','w',encoding='utf-8') as f:
    json.dump({
        'gerado_em': gerado_em,
        'periodo_inicio': f'01/01/{ano_atual}',
        'periodo_fim': hoje_str,
        'total_registros': len(recebimentos),
        'valor_total': total_recebido,
        'recebimentos': recebimentos,
    }, f, ensure_ascii=False, indent=2)

# ─── SALVAR EXCEL (3 abas) ────────────────────────────────────────
print('Salvando Excel...')
wb = Workbook()
roxo       = PatternFill(start_color='5B52E8', end_color='5B52E8', fill_type='solid')
roxo_claro = PatternFill(start_color='F4F6FB', end_color='F4F6FB', fill_type='solid')
verde      = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
verde_claro= PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
azul       = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
azul_claro = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
branco_ft  = Font(color='FFFFFF', bold=True, size=11, name='Arial')
borda = Border(
    left=Side(style='thin',color='E2E6F0'), right=Side(style='thin',color='E2E6F0'),
    top=Side(style='thin',color='E2E6F0'),  bottom=Side(style='thin',color='E2E6F0')
)

def escrever_aba(ws, headers, rows, fill_h, fill_alt, col_valor=None, larguras=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill=fill_h; cell.font=branco_ft
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border=borda
    ws.row_dimensions[1].height = 28
    prev = None; usar_alt = False
    for i, r in enumerate(rows, 2):
        chave = r.get('CNPJ/CPF') or r.get('CNPJ') or i
        if chave != prev: usar_alt = not usar_alt; prev = chave
        for col, h in enumerate(headers, 1):
            v = r.get(h,'')
            cell = ws.cell(row=i, column=col, value=v)
            cell.border=borda; cell.font=Font(name='Arial',size=10)
            if usar_alt: cell.fill=fill_alt
            if col_valor and col == col_valor: cell.number_format='R$ #,##0.00'
    if larguras:
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes='A2'
    ws.auto_filter.ref=f'A1:{get_column_letter(len(headers))}1'

# Aba 1: Inadimplentes
ws1 = wb.active; ws1.title='Inadimplentes'
h1 = ['Razao Social','Nome Fantasia','CNPJ/CPF','Email','Telefone','Cidade','Estado',
      'Regiao','Tags Regiao','Outras Tags','Diretora Regional','Gestora','Modelo de Negocio',
      'Distrato','Qtd Titulos','Valor Atrasado (R$)','Dias em Atraso','Vencimento Mais Antigo',
      'Ano Vencimento','Status','Documento','Categoria','Vencimento','Valor Titulo (R$)','Codigo Omie']
escrever_aba(ws1, h1, rows_inad_xlsx, roxo, roxo_claro, col_valor=16,
             larguras=[35,28,18,30,14,18,7,20,20,28,20,20,18,10,10,18,12,18,8,12,18,20,14,16,14])

# Aba 2: Recebidos
ws2 = wb.create_sheet('Recebidos')
h2 = ['Data','Conciliacao','Cliente','CNPJ','Categoria','Conta Corrente','Documento','Tipo Doc','Situacao','Vendedor','Valor']
escrever_aba(ws2, h2, recebimentos, verde, verde_claro, col_valor=11,
             larguras=[13,13,35,18,22,22,18,12,16,20,14])

# Aba 3: A Receber
ws3 = wb.create_sheet('A Receber')
h3 = ['Razao Social','Nome Fantasia','CNPJ/CPF','Email','Cidade','Estado','Regiao',
      'Gestora','Modelo de Negocio','Distrato','Documento','Categoria','Vencimento','Ano','Valor (R$)','Codigo Omie']
escrever_aba(ws3, h3, rows_avencer, azul, azul_claro, col_valor=15,
             larguras=[35,28,18,30,18,7,20,20,18,10,18,20,14,8,16,14])

wb.save('clubcasa-completo.xlsx')
print('  clubcasa-completo.xlsx salvo')

# ─── GOOGLE SHEETS (4 abas) ───────────────────────────────────────
print('\nAtualizando Google Sheets...')
try:
    gc = get_gc()
    ss = gc.open_by_key(SHEET_ID)

    # Aba Inadimplentes
    ws = garantir_aba(ss, 'Inadimplentes', h1)
    if rows_inad_xlsx:
        ws.append_rows([[r.get(h,'') for h in h1] for r in rows_inad_xlsx], value_input_option='RAW')
    print(f'  Inadimplentes: {len(rows_inad_xlsx)} linhas')

    # Aba Recebidos
    ws = garantir_aba(ss, 'Recebidos', h2)
    if recebimentos:
        ws.append_rows([[r.get(h,'') for h in h2] for r in recebimentos], value_input_option='RAW')
    print(f'  Recebidos: {len(recebimentos)} linhas')

    # Aba A Receber
    ws = garantir_aba(ss, 'A Receber', h3)
    if rows_avencer:
        ws.append_rows([[r.get(h,'') for h in h3] for r in rows_avencer], value_input_option='RAW')
    print(f'  A Receber: {len(rows_avencer)} linhas')

    # Aba Dashboard (KPIs)
    ws_dash = garantir_aba(ss, 'Dashboard', ['Metrica','Valor','Atualizado em'])
    kpis = [
        ['Total Inadimplentes (clientes)', len(rows_inad_json), gerado_em],
        ['Total Titulos em Atraso', total_tit, gerado_em],
        ['Valor Total Inadimplente (R$)', total_valor, gerado_em],
        ['Valor em Distrato (R$)', total_distrato, gerado_em],
        ['Total Recebido 2026 (R$)', total_recebido, gerado_em],
        ['Total A Receber (R$)', total_avencer_valor, gerado_em],
        ['Qtd Lancamentos Recebidos', len(recebimentos), gerado_em],
        ['Qtd Titulos A Receber', len(rows_avencer), gerado_em],
    ]
    ws_dash.append_rows(kpis, value_input_option='RAW')
    print(f'  Dashboard: {len(kpis)} KPIs')

except Exception as e:
    print(f'  ERRO Google Sheets: {e}')

# ─── GIT PUSH (Vercel) ────────────────────────────────────────────
print('\nPublicando no Vercel...')
try:
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    subprocess.run(['git','add','inadimplentes.json','recebimentos.json','clubcasa-completo.xlsx','index.html'], check=True)
    subprocess.run(['git','commit','-m',f'auto: atualização diária {gerado_em}'], check=True)
    subprocess.run(['git','push'], check=True)
    print('  Vercel atualizado!')
except Exception as e:
    print(f'  ERRO git push: {e}')

print(f'\n{"="*60}')
print(f'CONCLUÍDO — {gerado_em}')
print(f'Inadimplentes:  {len(rows_inad_json)} clientes | R$ {total_valor:,.2f}')
print(f'Distrato:       R$ {total_distrato:,.2f}')
print(f'Recebidos:      {len(recebimentos)} lançamentos | R$ {total_recebido:,.2f}')
print(f'A Receber:      {len(rows_avencer)} títulos | R$ {total_avencer_valor:,.2f}')
print(f'{"="*60}')
