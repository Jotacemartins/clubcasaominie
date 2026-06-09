import requests, time, unicodedata
from datetime import datetime, date
import gspread
from google.oauth2.service_account import Credentials

# ================================================================
# CONFIGURAÇÕES - edite aqui
# ================================================================
APP_KEY    = '1576318757013'
APP_SECRET = '2d96d060ad7abe5dd82c4d214f3a9de8'
CREDENTIALS_FILE = 'credentials.json'   # arquivo JSON da service account Google
SPREADSHEET_ID   = '1FzzxksfyE0K7PHMGGuLxrAFL35mTtag80Avr97B1Cjw'
SHEET_NAME       = 'Página1'

OMIE_CLIENTES   = 'https://app.omie.com.br/api/v1/geral/clientes/'
OMIE_CONTAS     = 'https://app.omie.com.br/api/v1/financas/contareceber/'
OMIE_VENDEDORES = 'https://app.omie.com.br/api/v1/geral/vendedores/'
OMIE_CATEGORIAS = 'https://app.omie.com.br/api/v1/geral/categorias/'

CABECALHOS = [
    'Razao Social', 'Nome Fantasia', 'CNPJ/CPF', 'Email', 'Telefone',
    'Cidade', 'Estado', 'Regiao', 'Diretora Regional', 'Gestora',
    'Modelo de Negocio', 'Distrato', 'Qtd Titulos', 'Valor Atrasado (R$)',
    'Dias em Atraso', 'Vencimento Mais Antigo', 'Status', 'Documento',
    'Categoria', 'Vencimento', 'Valor', 'Codigo Omie', 'Tags'
]

TAGS_REGIAO = {
    'regiao abc', 'regiao alphaville', 'regiao litoral paulista',
    'regiao oeste e sul', 'regiao sorocaba', 'sao paulo',
    'zona leste', 'zona norte', 'guarulhos', 'litoral sul',
    'alagoas', 'belo horizonte', 'brasilia', 'curitiba',
    'rio de janeiro', 'porto alegre', 'fortaleza', 'recife',
    'salvador', 'manaus', 'belem', 'goiania', 'florianopolis',
    'campo grande', 'maceio', 'natal', 'teresina',
    'santa catarina', 'parana', 'noroeste paulista',
    'interior sp', 'grande sp', 'litoral', 'sul', 'minas gerais'
}

TAGS_MODELO = {
    'associados pro', 'associados', 'design',
    'cessao de direitos', 'associado pro', 'associado'
}

# ================================================================
# FUNÇÕES AUXILIARES
# ================================================================
def norm(s):
    if not s: return ''
    s = unicodedata.normalize('NFD', str(s))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower().strip()

def eh2026(data_br):
    if not data_br: return False
    partes = str(data_br).split('/')
    return len(partes) == 3 and int(partes[2]) == 2026

def calcular_dias(data_br):
    if not data_br: return 0
    try:
        d, m, y = data_br.split('/')
        venc = date(int(y), int(m), int(d))
        return max(0, (date.today() - venc).days)
    except:
        return 0

def doc_titulo(t):
    nf = t.get('numero_documento_fiscal')
    if nf and str(nf).strip(): return f'NF {nf}'
    ndoc = t.get('numero_documento')
    if ndoc and str(ndoc).strip(): return str(ndoc)
    cod = t.get('codigo_lancamento_omie')
    return f'ID {cod}' if cod else 'N/I'

def omie_post(url, body):
    for i in range(3):
        try:
            r = requests.post(url, json=body, timeout=30)
            return r.json()
        except:
            if i == 2: raise
            time.sleep(2 ** i)

# ================================================================
# 1. VENDEDORES
# ================================================================
print('[1/4] Buscando vendedores...')
vendedores = {}
p, tp = 1, 1
while p <= tp:
    r = omie_post(OMIE_VENDEDORES, {
        'call': 'ListarVendedores', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
        'param': [{'pagina': p, 'registros_por_pagina': 100}]
    })
    if not r or not r.get('cadastro'): break
    tp = r.get('total_de_paginas', 1)
    for v in r['cadastro']:
        if v.get('codigo') and v.get('nome'):
            vendedores[str(v['codigo'])] = v['nome']
    p += 1
    time.sleep(0.2)
print(f'  {len(vendedores)} vendedores')

# ================================================================
# 2. CATEGORIAS
# ================================================================
print('[2/4] Buscando categorias...')
categorias = {}
p, tp = 1, 1
while p <= tp:
    r = omie_post(OMIE_CATEGORIAS, {
        'call': 'ListarCategorias', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
        'param': [{'pagina': p, 'registros_por_pagina': 100}]
    })
    if not r or not r.get('categoria_cadastro'): break
    tp = r.get('total_de_paginas', 1)
    for c in r['categoria_cadastro']:
        categorias[c['codigo']] = c.get('descricao') or c['codigo']
    p += 1
    time.sleep(0.2)
print(f'  {len(categorias)} categorias')

# ================================================================
# 3. CLIENTES
# ================================================================
print('[3/4] Buscando clientes...')
clientes = {}
p, tp = 1, 1
while p <= tp:
    try:
        r = omie_post(OMIE_CLIENTES, {
            'call': 'ListarClientes', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
            'param': [{'pagina': p, 'registros_por_pagina': 100,
                       'apenas_importado_api': 'N', 'exibir_caracteristicas': 'S'}]
        })
    except:
        p += 1
        time.sleep(0.5)
        continue

    if not r or not r.get('clientes_cadastro'): break
    tp = r.get('total_de_paginas', 1)

    for c in r['clientes_cadastro']:
        cnpj = c.get('cnpj_cpf', '') or ''
        nome = c.get('razao_social', '') or ''
        if not cnpj or cnpj == '000.000.000-00' or not nome.strip() or nome == 'Cliente Consumidor':
            continue

        raw_tags = [t.get('tag', '') for t in c.get('tags', [])]
        nd_tags  = [norm(t) for t in raw_tags]

        # Exclui só quem é exclusivamente fornecedor
        if 'fornecedor' in nd_tags and 'cliente' not in nd_tags:
            continue

        diretora = gestora = regiao_c = modelo_c = status_c = ''
        for carac in c.get('caracteristicas', []):
            campo    = norm(carac.get('campo', ''))
            conteudo = str(carac.get('conteudo', '')).strip()
            if (('diretor' in campo or 'franquia' in campo or 'regional' in campo)
                    and 'regiao' not in campo):
                diretora = conteudo
            if 'gestor' in campo:       gestora  = conteudo
            elif 'regiao' in campo:     regiao_c = conteudo
            elif 'modelo' in campo:     modelo_c = conteudo
            elif 'status' in campo:     status_c = norm(conteudo)

        regiao_tag = next((raw_tags[i] for i, t in enumerate(nd_tags) if t in TAGS_REGIAO), '')
        modelo_tag = next((raw_tags[i] for i, t in enumerate(nd_tags) if t in TAGS_MODELO), '')
        distrato   = 'Sim' if (status_c == 'distrato' or 'distrato' in nd_tags) else 'Nao'

        cod_vend = str((c.get('recomendacoes') or {}).get('codigo_vendedor', '') or '')
        if not gestora and cod_vend:
            gestora = vendedores.get(cod_vend, '')

        clientes[str(c['codigo_cliente_omie'])] = {
            'razao_social':      nome,
            'nome_fantasia':     c.get('nome_fantasia', '') or '',
            'cnpj_cpf':          cnpj,
            'email':             (c.get('email', '') or '').split(',')[0].strip(),
            'telefone':          c.get('telefone1_numero', '') or '',
            'cidade':            c.get('cidade', '') or '',
            'estado':            c.get('estado', '') or '',
            'regiao':            (regiao_c or regiao_tag) or 'Sem Regiao',
            'diretora_regional': diretora or 'Nao Informada',
            'gestora':           gestora  or 'Nao Informada',
            'modelo':            (modelo_c or modelo_tag) or 'Sem Modelo',
            'distrato':          distrato,
            'tags':              ', '.join(raw_tags),
            'codigo_omie':       c['codigo_cliente_omie'],
        }

    print(f'  Pag {p}/{tp} - {len(clientes)} clientes')
    p += 1
    time.sleep(0.25)

# ================================================================
# 4. TÍTULOS
# ================================================================
print('[4/4] Buscando títulos...')
linhas = []

STATUS_MAP = [
    {'filtro': 'ATRASADO',  'label': 'Inadimplente', 'so2026': False, 'data_de': None,         'data_ate': None},
    {'filtro': 'AVENCER',   'label': 'A Vencer',     'so2026': False, 'data_de': None,         'data_ate': None},
    {'filtro': 'PAGO',      'label': 'Pago',          'so2026': True,  'data_de': '01/01/2026', 'data_ate': '31/12/2026'},
    {'filtro': 'LIQUIDADO', 'label': 'Pago',          'so2026': True,  'data_de': '01/01/2026', 'data_ate': '31/12/2026'},
]

for st in STATUS_MAP:
    p, tp = 1, 1
    count = 0
    erros = 0

    while p <= tp:
        try:
            r = omie_post(OMIE_CONTAS, {
                'call': 'ListarContasReceber', 'app_key': APP_KEY, 'app_secret': APP_SECRET,
                'param': [{'pagina': p, 'registros_por_pagina': 100,
                           'filtrar_por_status': st['filtro'],
                           'filtrar_apenas_titulos_em_aberto': 'N'}]
            })
        except:
            erros += 1
            if erros >= 3: break
            time.sleep(1)
            continue

        if not r: break

        if r.get('faultstring'):
            erros += 1
            print(f'  Aviso pag {p} [{st["filtro"]}]: {r["faultstring"]}')
            if erros >= 5: break
            p += 1
            time.sleep(1)
            continue

        erros = 0
        tp = r.get('total_de_paginas', 1)
        achou_na_pag = False

        for t in (r.get('conta_receber_cadastro') or r.get('lista_contareceber') or []):
            venc = t.get('data_vencimento') or t.get('dDtVenc') or ''

            # Pagos: só 2026 | Atrasados e A Vencer: todos
            if st['so2026'] and not eh2026(venc):
                continue
            achou_na_pag = True

            cod = str(t.get('codigo_cliente_fornecedor', ''))
            cl  = clientes.get(cod)
            if not cl: continue

            valor = float(t.get('valor_documento') or 0)

            linhas.append([
                cl['razao_social'],
                cl['nome_fantasia'],
                cl['cnpj_cpf'],
                cl['email'],
                cl['telefone'],
                cl['cidade'],
                cl['estado'],
                cl['regiao'],
                cl['diretora_regional'],
                cl['gestora'],
                cl['modelo'],
                cl['distrato'],
                1,                                                          # Qtd Titulos
                valor if st['filtro'] == 'ATRASADO' else 0,                # Valor Atrasado
                calcular_dias(venc) if st['filtro'] == 'ATRASADO' else 0,  # Dias em Atraso
                venc,                                                       # Vencimento Mais Antigo
                st['label'],
                doc_titulo(t),
                categorias.get(t.get('codigo_categoria', ''), 'Sem Categoria'),
                venc,
                valor,
                cl['codigo_omie'],
                cl['tags'],
            ])
            count += 1

        print(f'  [{st["filtro"]}] Pag {p}/{tp} - {count} títulos')

        # Para pagos: se não achou nada de 2026 em 10 páginas seguidas, desiste
        if st['so2026'] and not achou_na_pag:
            pags_sem_2026 = getattr(st, '_pags_sem_2026', 0) + 1
            st['_pags_sem_2026'] = pags_sem_2026
            if pags_sem_2026 >= 10:
                print(f'  Sem dados de 2026 por 10 paginas, encerrando {st["filtro"]}')
                break
        else:
            st['_pags_sem_2026'] = 0

        p += 1
        time.sleep(0.3)

    print(f'  {st["label"]} ({st["filtro"]}): {count} títulos')

print(f'\nTotal: {len(linhas)} linhas geradas')


# ================================================================
# 5. SALVAR XLSX LOCAL
# ================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print(f'\nSalvando arquivo local...')
wb = Workbook()
ws = wb.active
ws.title = 'Dados'

roxo   = PatternFill(start_color='5B52E8', end_color='5B52E8', fill_type='solid')
branco = Font(color='FFFFFF', bold=True)

for col, h in enumerate(CABECALHOS, start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = roxo
    cell.font = branco
    cell.alignment = Alignment(horizontal='center')

for i, linha in enumerate(linhas, start=2):
    for col, val in enumerate(linha, start=1):
        ws.cell(row=i, column=col, value=val)

larguras = [35, 30, 18, 30, 15, 20, 8, 22, 20, 20, 20, 10, 10, 18, 12, 18, 14, 20, 20, 18, 12, 14, 30]
for i, w in enumerate(larguras, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A2'
nome_arquivo = 'clubcasa-financeiro.xlsx'
wb.save(nome_arquivo)

print(f'\n✅ CONCLUÍDO!')
print(f'Total de linhas: {len(linhas)}')
print(f'Arquivo salvo: {nome_arquivo}')